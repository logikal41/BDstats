#! python3
# statsAnalysis.py -- Program used to pull data from test log csv file, analyse, and export to excel test doc

"""   Import Modules   """
import csv, sys, numpy, openpyxl, copy
import scipy.stats as stats  #only need t.ppf , import this function alone
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.chart import ScatterChart, Reference, Series

#Import testomatic file. Open file and read it into list "testData" (with error detection)
fileName = input('Please enter the file name: ')
try:
	testomaticFile = open(fileName+'.csv')
	testomaticReader = csv.reader(testomaticFile)
	testDataList = list(testomaticReader)
except FileNotFoundError:
	print('Invalid File Name!')
	sys.exit()

"""   Global Variables   """

rawData = []  #Raw data from original file
testData = [] #List used for filtered data
manip = 'y'  #Used to trigger data filtering
filtersUsed = []  #Filter tracking list
units = '' #This will track what units are requested
selectedConfInt = '' #Used to track the selected confidence interval

"""   DEFINITIONS   """

#Read data point into a dictionary and insert into the testData list
#Keys with no value show up as an empty string ('')
def populateTestData():
	print('Retrieving Data...')
	for row in range(16,len(testDataList)):
		#Pull the individual data and assign to local variable
		testRun = testDataList[row][0]
		procedure = testDataList[row][2]
		configuration = testDataList[row][3]
		peakLoad = testDataList[row][4]
		failureType = testDataList[row][7]
		failureNotes = testDataList[row][8]
		color = testDataList[row][9]
		size = testDataList[row][10]
		#Initialize data point dictionary (local variable)
		dataPoint = {'test run': '', 'procedure': '', 'configuration': '', 'peak load': '', \
		'failure type': '', 'failure notes': '', 'color': '', 'size': ''}
		#Populate data point dictionary 
		dataPoint['test run'] = int(testRun)
		dataPoint['procedure'] = str(procedure).lower()
		dataPoint['configuration'] = str(configuration).lower()
		dataPoint['peak load'] = float(peakLoad)
		dataPoint['failure type'] = str(failureType).lower()
		dataPoint['failure notes'] = str(failureNotes).lower()
		dataPoint['color'] = str(color).lower()
		dataPoint['size'] = str(size).lower()
		#Append data point dictionary to testData list (global variable)
		testData.append(dataPoint)
	print('Done.')
	print('') #Print a blank line for readability

#Create a copy of the original data for reference purposes
#This is necessary to remove any dictionary/list python references 
def duplicateData(dataSet):
	data = []
	for i in range(len(dataSet)):
		data.append(copy.copy(dataSet[i]))
	return data

#Print out the filter key list excluding 'peak load' and 'test run'
def printFilterKeys(keyList):
	for key in keyList.keys():
		if key == 'peak load':
			continue
		elif key == 'test run':
			continue
		else:
			print(key)
	print('') #Add a blank line for readability

#Function for searching a list for filters
def searchList(dataSet, filterKey, filterValue):
	for i in range(len(dataSet)):
		if dataSet[i][filterKey] != filterValue:
			return i

#Perform data filtering on testData list
def filterData():
	global filtersUsed
	filterSelection = [] #local variable
	print('')
	print('Filter categories:')
	printFilterKeys(testData[0]) #testData[0] is arbitrary to get key information
	filterKey = input('What would you like to filter by? (hit enter to exit) ').lower()  #local variable
	#Error Detection
	while True:
		if filterKey == '':
			return
		try:
			instances = [] #local variable
			for i in range(len(testData)):
				if str(testData[i][filterKey]).lower() not in instances:
					instances.append(str(testData[i][filterKey]).lower())
			break
		except KeyError:
			print('')
			print('Invalid answer.')
			print('')
			filterKey = input('What would you like to filter by? (hit enter to exit) ').lower()  #local variable
	#Append each filter used for filter tracking and print out
	filterSelection.append(filterKey)
	print('')
	print('Filter categories:')
	print('\n'.join(instances))
	print('')
	filterValue = input('What would you like to filter by? (hit enter to exit) ').lower()
	#Error Detection
	while True:
		if filterValue == '':
			return
		elif filterValue not in instances:
			print('')
			print('Invalid answer.')
			print('')
			filterValue = input('What would you like to filter by? (hit enter to exit) ').lower()  #local variable
		else:
			break
	#Append each filter used for filter tracking and print out
	filterSelection.append(filterValue)
	#Remove any dictionaries that dont contain the filter from the list
	while True:
		activeFilter = searchList(testData, filterKey, filterValue)
		if activeFilter != None:
			del testData[activeFilter]
		else:
			break
	#Remove filtered key so it cant be used again  
	for j in range(len(testData)):
		del testData[j][filterKey]
	print('')
	#Append filter selections to the global variable filtersUsed
	filtersUsed.append(filterSelection)

#Asks to continue filtering
def askToFilter():
	global manip  #do we need this global variable??????
	manip = input('Would you like to filter the data? (Y/N) ').lower()
	print('')
	while True:
		if manip == 'y':
			filterData()
			print('')
			return manip
		elif manip == 'n':
			print('No more filters applied.')
			print('')
			return manip
		else:
			print('Invalid answer.')
			print('')
			manip = input('Would you like to filter the data? (Y/N) ').lower()
			print('')

#Convert testData from pounds to kilonewtons
#For use with list of dictionaries
def convertToMetric(dataSet):
	for i in range(len(dataSet)):
		dataSet[i]['peak load'] = float(dataSet[i]['peak load'])/224.8
	return dataSet

#Build CDF value list
def cdf(n):
	cdf = [1/(n+1)]
	for i in range(1,n):
		nextCDF = cdf[i-1]+(1/(n+1))
		cdf.append(nextCDF)
	return cdf

#Sort the data for normality checks
#For use with list of dictionaries
def dataSort(dataSet):
	peakLoads = [] #local variable
	for i in range(len(dataSet)):
		peakLoads.append(dataSet[i]['peak load'])
	peakLoads.sort()
	return peakLoads

#Get average function
#For use with dictionaries
def getAverage(dataSet):
	sum = 0
	for i in range(len(dataSet)):
		sum += float(dataSet[i]['peak load'])
	average = sum / len(dataSet)
	return average

#Get Standard Deviation
#For use with dictionaries
def getStdDev(dataSet):
	sum = 0
	for i in range(len(dataSet)):
		sum += (float(dataSet[i]['peak load']) - getAverage(dataSet))**2
	try:
		stdDev = (sum / (len(dataSet)-1)) ** (1/2)
	except ZeroDivisionError:
		stdDev = 'n/a'
	return stdDev

#Get lower 3 sigma value
#For use with dictionaries
def threeSigma(dataSet):
	threeSigma = getAverage(dataSet) - (3 * getStdDev(dataSet))
	return threeSigma

#Confidence Interval Calculation
def confidenceInterval(dataSet):
	global selectedConfInt
	intervalChoices = ['85', '90', '95']
	interval = input('Choose your Confidence Interval (85, 90, or 95): ') 
	while True:
		if interval not in intervalChoices:
			print('')
			print('Invalid interval.')
			print('')
			interval = input('Choose your Confidence Interval (85, 90, or 95): ')
			print('')
		else:
			selectedConfInt = interval + '%'
			break
	interval = (1 - (int(interval)/100))/2
	confidenceInterval = stats.t.ppf(1-interval, (len(dataSet)-1))
	return confidenceInterval

#Get key values
def getKeys(dataSet):
	keys = []
	for key in dataSet[0].keys():    #dataSet[0] is an arbitrary dictionary to get the keys
		keys.append(key)
	return keys

#Populate and format data into excel
def populateTestDoc(sheet, dataSet, startRow, unitSelect):
	#Heading Font (set here)
	headingFont = Font(bold=True, size=12)
	#print all of the labels
	keys = getKeys(dataSet)
	#update peak load key with appropriate units
	updateUnits = 'peak load ' + '(' + str(unitSelect) + ')'
	keys[keys.index('peak load')] = updateUnits
	for i in range(len(keys)):
		sheet.cell(row=startRow, column=i+1).value = keys[i]
	#Initialize column widths widths for formating
	col_widths = []
	for i in range(len(keys)):
		col_widths.append(len(keys[i]))
	#revert keys back
	keys[keys.index(updateUnits)] = 'peak load'
	#print data
	for row in range(len(dataSet)):
		for col in range(len(keys)):
			sheet.cell(row=(startRow+1)+row, column=1+col).value = dataSet[row][keys[col]]   
	#Format the table headers
	sheet.row_dimensions[1].height = 22
	for i in range(len(keys)):
		sheet.cell(row=startRow, column=i+1).font = headingFont
		sheet.cell(row=startRow, column=i+1).alignment = Alignment(vertical='center', horizontal='center')
	#Iterate over dataSet to update column widths and center align all cells
	for col in range(len(keys)):
		for row in range(len(dataSet)):
			sheet.cell(row=row+(startRow+1), column=col+1).alignment = Alignment(horizontal='center')
			if len(str(dataSet[row][keys[col]])) > col_widths[col]:
				col_widths[col] = len(str(dataSet[row][keys[col]]))
	#Edit column widths
	for i in range(len(col_widths)):
		sheet.column_dimensions[get_column_letter(i+1)].width = col_widths[i] + 2 #Add 2 char buffer to width 

#Print things to an excel spreadsheet for test documents
#For use with dictionary list
def writeTestDoc():
	global units
	newFileName = input('What would you like to name the test document? ')
	newFileName = newFileName + '.xlsx'
	print('')
	print('Writing to ' + newFileName + '...')
	wb = openpyxl.load_workbook('testDoc.xlsx')  #make sure testDoc.xlsx is in the root folder
	sheet1 = wb.get_sheet_by_name('Raw Data')
	sheet2 = wb.get_sheet_by_name('Statistics')
	#print all of the raw data to sheet 1 (raw data)
	populateTestDoc(sheet1, rawData, 1, 'lbf')
	#Merge cells and Print all the filters that have been applied
	sheet2.merge_cells('A2:F2')
	sheet2.cell(row=1, column=1).value = 'Filters Applied:'
	sheet2.cell(row=1, column=1).font = Font(bold=True, italic=True)
	sheet2['A2'] = str(filtersUsed)
	#print all of the testData to sheet 2 (statistics)
	populateTestDoc(sheet2, testData, 4, units)
	#Format Column Width for Stats Data
	sheet2.column_dimensions['A'].width = 23.5 
	#Print number of samples (row starts 2 after the input value for populateTestDoc)
	sheet2.cell(row= (len(testData)+5), column= 1).value = 'n:'
	sheet2.cell(row= (len(testData)+5), column= 1).alignment = Alignment(horizontal='right')
	sheet2.cell(row= (len(testData)+5), column= 2).value = len(testData)
	sheet2.cell(row= (len(testData)+5), column= 2).alignment = Alignment(horizontal='center')   
	#Print Average 
	sheet2.cell(row= (len(testData)+6), column= 1).value = 'Average:'
	sheet2.cell(row= (len(testData)+6), column= 1).alignment = Alignment(horizontal='right')
	sheet2.cell(row= (len(testData)+6), column= 2).value = average
	sheet2.cell(row= (len(testData)+6), column= 2).alignment = Alignment(horizontal='center')
	sheet2.cell(row= (len(testData)+6), column= 2).number_format = '0.0000'  
	#Print Standard Deviation
	sheet2.cell(row= (len(testData)+7), column= 1).value = 'Standard Deviation:'
	sheet2.cell(row= (len(testData)+7), column= 1).alignment = Alignment(horizontal='right')
	sheet2.cell(row= (len(testData)+7), column= 2).value = stdDev
	sheet2.cell(row= (len(testData)+7), column= 2).alignment = Alignment(horizontal='center')
	sheet2.cell(row= (len(testData)+7), column= 2).number_format = '0.0000' 
	#Print Variance
	sheet2.cell(row= (len(testData)+8), column= 1).value = 'Variance:'
	sheet2.cell(row= (len(testData)+8), column= 1).alignment = Alignment(horizontal='right')
	sheet2.cell(row= (len(testData)+8), column= 2).value = stdDev ** 2
	sheet2.cell(row= (len(testData)+8), column= 2).alignment = Alignment(horizontal='center')
	sheet2.cell(row= (len(testData)+8), column= 2).number_format = '0.0000'    
	#Print Three Sigma
	sheet2.cell(row= (len(testData)+9), column= 1).value = 'Lower Three Sigma:'
	sheet2.cell(row= (len(testData)+9), column= 1).alignment = Alignment(horizontal='right')
	sheet2.cell(row= (len(testData)+9), column= 2).value = threeSigma
	sheet2.cell(row= (len(testData)+9), column= 2).alignment = Alignment(horizontal='center')
	sheet2.cell(row= (len(testData)+9), column= 2).number_format = '0.0000'    
	#Print Confidence Interval
	sheet2.cell(row= (len(testData)+10), column= 1).value = selectedConfInt + ' Confidence Interval:'
	sheet2.cell(row= (len(testData)+10), column= 1).alignment = Alignment(horizontal='right')
	sheet2.cell(row= (len(testData)+10), column= 2).value = confidenceIntervalValue
	sheet2.cell(row= (len(testData)+10), column= 2).alignment = Alignment(horizontal='center')
	sheet2.cell(row= (len(testData)+10), column= 2).number_format = '0.0000'   
	#Index for end of filtered data analysis
	filteredDataEnd = len(testData)+10
	#Print QQ info 
	if len(testData) > 1:
		printQQ(sheet2, testData, (filteredDataEnd+2), 1) 
		printNormalCurve(sheet2, dataSort(testData), filteredDataEnd+len(testData)+9, 1)
	else:
		sheet2.cell(row=(filteredDataEnd+2), column=1).value = 'NOTE: Can not evaluate normality when n = 1'
		sheet2.cell(row=(filteredDataEnd+2), column=1).font = Font(bold=True)
	#Save and close file
	wb.save(newFileName)
	print('Done.') 

#print QQ data and plot
def printQQ(sheet, dataSet, startRow, startCol):
	#Print Heading
	selectedCells = get_column_letter(startCol) + str(startRow) + ':' + get_column_letter(startCol+3) + str(startRow)
	sheet.merge_cells(selectedCells)
	sheet.cell(row=startRow, column=startCol).value = 'Q-Q Plot Data'
	sheet.cell(row=startRow, column=startCol).font = Font(bold=True)
	sheet.cell(row=startRow, column=startCol).alignment = Alignment(horizontal='center')
	#Print actual data list column
	sheet.cell(row=startRow+1, column=startCol).value = 'Actual Data ' + '(' + units + ')'
	sheet.cell(row=startRow+1, column=startCol).alignment = Alignment(horizontal='center')
	sheet.cell(row=startRow+1, column=startCol).font = Font(bold=True) 
	sortedList = dataSort(dataSet)
	for i in range(2,len(sortedList)+2):
		sheet.cell(row=startRow+i, column=startCol).value = sortedList[i-2]
		sheet.cell(row=startRow+i, column=startCol).alignment = Alignment(horizontal='center')
		sheet.cell(row=startRow+i, column=startCol).number_format = '0.0000' 
	#Print CDF values
	cdfData = cdf(len(sortedList))
	sheet.cell(row=startRow+1, column=startCol+1).value = 'CDF Value'
	sheet.cell(row=startRow+1, column=startCol+1).alignment = Alignment(horizontal='center')
	sheet.cell(row=startRow+1, column=startCol+1).font = Font(bold=True) 
	for i in range(2,len(cdfData)+2):
		sheet.cell(row=startRow+i, column=startCol+1).value = cdfData[i-2]
		sheet.cell(row=startRow+i, column=startCol+1).alignment = Alignment(horizontal='center')
		sheet.cell(row=startRow+i, column=startCol+1).number_format = '0.0000' 
	#Print Expected Values
	average = getAverage(dataSet)
	stdDev = getStdDev(dataSet)
	sheet.cell(row=startRow+1, column=startCol+2).value = 'Expected Value'
	sheet.cell(row=startRow+1, column=startCol+2).alignment = Alignment(horizontal='center')
	sheet.cell(row=startRow+1, column=startCol+2).font = Font(bold=True) 
	for i in range(2,len(cdfData)+2):
		sheet.cell(row=startRow+i, column=startCol+2).value = stats.norm.ppf(cdfData[i-2], loc = average, scale = stdDev)
		sheet.cell(row=startRow+i, column=startCol+2).alignment = Alignment(horizontal='center')
		sheet.cell(row=startRow+i, column=startCol+2).number_format = '0.0000' 
	#Adjust the column width if necessary
	expectedValueWidth = sheet.column_dimensions[get_column_letter(startCol+2)].width
	if expectedValueWidth < 15:
		sheet.column_dimensions[get_column_letter(startCol+2)].width = 15
	#Print z-score
	sheet.cell(row=startRow+1, column=startCol+3).value = 'z Score'
	sheet.cell(row=startRow+1, column=startCol+3).alignment = Alignment(horizontal='center')
	sheet.cell(row=startRow+1, column=startCol+3).font = Font(bold=True) 
	for i in range(2,len(cdfData)+2):
		sheet.cell(row=startRow+i, column=startCol+3).value = stats.norm.ppf(cdfData[i-2])
		sheet.cell(row=startRow+i, column=startCol+3).alignment = Alignment(horizontal='center')
		sheet.cell(row=startRow+i, column=startCol+3).number_format = '0.0000' 
	#Print average
	sheet.cell(row=startRow+len(sortedList)+2, column= startCol).value = 'Average:'
	sheet.cell(row=startRow+len(sortedList)+2, column= startCol).alignment = Alignment(horizontal='right')
	sheet.cell(row=startRow+len(sortedList)+2, column= startCol+1).value = numpy.average(sortedList)
	sheet.cell(row=startRow+len(sortedList)+2, column= startCol+1).alignment = Alignment(horizontal='center')
	sheet.cell(row=startRow+len(sortedList)+2, column= startCol+1).number_format = '0.0000'  
	#Print median
	sheet.cell(row=startRow+len(sortedList)+3, column= startCol).value = 'Median:'
	sheet.cell(row=startRow+len(sortedList)+3, column= startCol).alignment = Alignment(horizontal='right')
	sheet.cell(row=startRow+len(sortedList)+3, column= startCol+1).value = numpy.median(sortedList)
	sheet.cell(row=startRow+len(sortedList)+3, column= startCol+1).alignment = Alignment(horizontal='center') 
	sheet.cell(row=startRow+len(sortedList)+3, column= startCol+1).number_format = '0.0000' 
	#Print skewness (corrected for statistical bias)
	sheet.cell(row=startRow+len(sortedList)+4, column= startCol).value = 'Skewness:'
	sheet.cell(row=startRow+len(sortedList)+4, column= startCol).alignment = Alignment(horizontal='right')
	sheet.cell(row=startRow+len(sortedList)+4, column= startCol+1).value = stats.skew(sortedList, bias=False) #corrected for statistical bias
	sheet.cell(row=startRow+len(sortedList)+4, column= startCol+1).alignment = Alignment(horizontal='center') 
	sheet.cell(row=startRow+len(sortedList)+4, column= startCol+1).number_format = '0.0000' 
	#Print Kurtosis (corrected for statistical bias)
	sheet.cell(row=startRow+len(sortedList)+5, column= startCol).value = 'Kurtosis:'
	sheet.cell(row=startRow+len(sortedList)+5, column= startCol).alignment = Alignment(horizontal='right')
	sheet.cell(row=startRow+len(sortedList)+5, column= startCol+1).value = stats.kurtosis(sortedList, bias=False) #corrected for statistical bias
	sheet.cell(row=startRow+len(sortedList)+5, column= startCol+1).alignment = Alignment(horizontal='center')
	sheet.cell(row=startRow+len(sortedList)+5, column= startCol+1).number_format = '0.0000' 
	#Index for QQ plot data end
	qqPlotEnd = len(sortedList)+5
	#QQ plot (Line Chart)
	chart = ScatterChart()
	chart.title = "Q-Q Plot"
	chart.style = 13
	chart.x_axis.title = 'z Score'
	chart.y_axis.title = 'Peak Load'
	xvalues = Reference(sheet, min_col=startCol+3, min_row=startRow+2, max_row=startRow+(len(cdfData)+2))
	for i in range(startCol, (startCol+3), 2):
		values = Reference(sheet, min_col=i, min_row=(startRow+1), max_row=(startRow+len(sortedList)+2))
		series = Series(values, xvalues, title_from_data=True)
		chart.series.append(series)
	#chart location
	chartAnchor =  get_column_letter(startCol+4) + str(startRow-5)
	sheet.add_chart(chart, chartAnchor)

#Print normal curve with confidence bounds
def printNormalCurve(sheet, sortedList, startRow, startCol):
	#Print Heading
	selectedCells1 = get_column_letter(startCol) + str(startRow) + ':' + get_column_letter(startCol+3) + str(startRow)
	sheet.merge_cells(selectedCells1)
	sheet.cell(row=startRow, column=startCol).value = 'Bell Curve Plot Data'
	sheet.cell(row=startRow, column=startCol).font = Font(bold=True)
	sheet.cell(row=startRow, column=startCol).alignment = Alignment(horizontal='center')
	#Print settings
	selectedCells2 = get_column_letter(startCol) + str(startRow+1) + ':' + get_column_letter(startCol+1) + str(startRow+1)
	sheet.merge_cells(selectedCells2)
	sheet.cell(row=startRow+1, column=startCol).value = 'Settings'
	sheet.cell(row=startRow+1, column=startCol).alignment = Alignment(horizontal='center')
	sheet.cell(row=startRow+1, column=startCol).font = Font(bold=True)
	#Z settings 
	zMin = -4 #initialize Z min setting
	zMax = 4 #initialize Z max setting
	sheet.cell(row=startRow+2, column= startCol).value = 'Z_min:'
	sheet.cell(row=startRow+2, column= startCol).alignment = Alignment(horizontal='right')
	sheet.cell(row=startRow+2, column= startCol+1).value = zMin #number of standard deviations
	sheet.cell(row=startRow+2, column= startCol+1).alignment = Alignment(horizontal='center')
	sheet.cell(row=startRow+3, column= startCol).value = 'Z_max:'
	sheet.cell(row=startRow+3, column= startCol).alignment = Alignment(horizontal='right')
	sheet.cell(row=startRow+3, column= startCol+1).value = zMax #number of standard deviations
	sheet.cell(row=startRow+3, column= startCol+1).alignment = Alignment(horizontal='center')
	#Confidence Bounds 
	sheet.cell(row=startRow+4, column= startCol).value = 'Lower Confidence Bound:'
	sheet.cell(row=startRow+4, column= startCol).alignment = Alignment(horizontal='right')
	lowerConfBound = average - confidenceInterval*stdDev #local variable
	sheet.cell(row=startRow+4, column= startCol+1).value = lowerConfBound
	sheet.cell(row=startRow+4, column= startCol+1).alignment = Alignment(horizontal='center')
	sheet.cell(row=startRow+4, column= startCol+1).number_format = '0.0000'
	sheet.cell(row=startRow+5, column= startCol).value = 'Upper Confidence Bound:'
	sheet.cell(row=startRow+5, column= startCol).alignment = Alignment(horizontal='right')
	upperConfBound = average + confidenceInterval*stdDev #local variable
	sheet.cell(row=startRow+5, column= startCol+1).value = upperConfBound
	sheet.cell(row=startRow+5, column= startCol+1).alignment = Alignment(horizontal='center')
	sheet.cell(row=startRow+5, column= startCol+1).number_format = '0.0000'
	#Curve Bounds 
	sheet.cell(row=startRow+6, column= startCol).value = 'Curve Min:'
	sheet.cell(row=startRow+6, column= startCol).alignment = Alignment(horizontal='right')
	sheet.cell(row=startRow+6, column= startCol+1).value = average - 5*stdDev
	sheet.cell(row=startRow+6, column= startCol+1).alignment = Alignment(horizontal='center')
	sheet.cell(row=startRow+6, column= startCol+1).number_format = '0.0000'
	sheet.cell(row=startRow+7, column= startCol).value = 'Curve Max:'
	sheet.cell(row=startRow+7, column= startCol).alignment = Alignment(horizontal='right')
	sheet.cell(row=startRow+7, column= startCol+1).value = average + 5*stdDev
	sheet.cell(row=startRow+7, column= startCol+1).alignment = Alignment(horizontal='center')
	sheet.cell(row=startRow+7, column= startCol+1).number_format = '0.0000'
	#Print Z column heading
	sheet.cell(row=startRow+8, column=startCol).value = 'Z'
	sheet.cell(row=startRow+8, column=startCol).alignment = Alignment(horizontal='center')
	sheet.cell(row=startRow+8, column=startCol).font = Font(bold=True)
	#Print X column heading
	sheet.cell(row=startRow+8, column=startCol+1).value = 'X'
	sheet.cell(row=startRow+8, column=startCol+1).alignment = Alignment(horizontal='center')
	sheet.cell(row=startRow+8, column=startCol+1).font = Font(bold=True)
	#Print f(x) column heading
	sheet.cell(row=startRow+8, column=startCol+2).value = 'f(x)'
	sheet.cell(row=startRow+8, column=startCol+2).alignment = Alignment(horizontal='center')
	sheet.cell(row=startRow+8, column=startCol+2).font = Font(bold=True)
	#Print Area column heading
	sheet.cell(row=startRow+8, column=startCol+3).value = 'Area'
	sheet.cell(row=startRow+8, column=startCol+3).alignment = Alignment(horizontal='center')
	sheet.cell(row=startRow+8, column=startCol+3).font = Font(bold=True)
	#Print data
	z = zMin #initialize counter 
	for i in range(9,len(sortedList)+9):
		#Z data
		sheet.cell(row=startRow+i, column=startCol).value = z
		sheet.cell(row=startRow+i, column=startCol).alignment = Alignment(horizontal='center')
		sheet.cell(row=startRow+i, column=startCol).number_format = '0.0000' 
		#X data
		x = z*stdDev + average
		sheet.cell(row=startRow+i, column=startCol+1).value = x
		sheet.cell(row=startRow+i, column=startCol+1).alignment = Alignment(horizontal='center')
		sheet.cell(row=startRow+i, column=startCol+1).number_format = '0.0000' 
		#f(x) data
		sheet.cell(row=startRow+i, column=startCol+2).value = stats.norm.pdf(x, loc=average, scale=stdDev) # NEED TO TURN OFF CUMULATIVE FUNCTION!!!
		sheet.cell(row=startRow+i, column=startCol+2).alignment = Alignment(horizontal='center')
		sheet.cell(row=startRow+i, column=startCol+2).number_format = '0.0000' 
		#Area data
		if x > upperConfBound or x < lowerConfBound:
			sheet.cell(row=startRow+i, column=startCol+3).value = ''
		else:
			sheet.cell(row=startRow+i, column=startCol+3).value = stats.norm.pdf(x, loc=average, scale=stdDev)
		sheet.cell(row=startRow+i, column=startCol+3).alignment = Alignment(horizontal='center')
		z = (zMax-zMin)/(len(sortedList)-1) + z
	#Bell curve. Scatter with smooth lines
	chart = ScatterChart()
	chart.title = "Normal Curve"
	chart.style = 13
	chart.x_axis.title = 'peak load (' + units +')' 
	chart.y_axis.title = 'f(x)'
	xvalues = Reference(sheet, min_col=startCol+1, min_row=startRow+9, max_row=startRow+(len(sortedList)+8))
	values1 = Reference(sheet, min_col=startCol+2, min_row=startRow+9, max_row=startRow+(len(sortedList)+8))
	series1 = Series(values1, xvalues)
	values2 = Reference(sheet, min_col=startCol+3, min_row=startRow+9, max_row=startRow+(len(sortedList)+8))
	series2 = Series(values2, xvalues)
	chart.series.append(series1)
	chart.series.append(series2)
	#set a buffer for x axis scale
	chart.x_axis.scaling.min = (sheet.cell(row=startRow+9, column=startCol+1).value) - (sheet.cell(row=startRow+9, column=startCol+1).value)*.10 
	chart.x_axis.scaling.max = (sheet.cell(row=(startRow+8+len(sortedList)), column=startCol+1).value) + \
	(sheet.cell(row=(startRow+8+len(sortedList)), column=startCol+1).value)*.10
	#chart location
	chartAnchor =  get_column_letter(startCol+4) + str(startRow-4)
	sheet.add_chart(chart, chartAnchor)


""" MAIN BODY OF CODE   """

#Populate the rawData list
populateTestData()
rawData = duplicateData(testData)  #remove python references to original dictionaries imported from csv file

#Apply filters if applicable
while manip == 'y': #This global variable has been pre-set to 'y'
	askToFilter()
	print('Filters Used: ' + str(filtersUsed))
	print('')

#Print out the final filtered testData (peak load) list 
print('Filtered Data (lbf): ')
for i in range(len(testData)):
	print(testData[i]['peak load'])
print('')
print('Filters Used: ' + str(filtersUsed))
print('')

#Ask what unit to report the information in
#Use this global to track what units are being requested for print outs. 
units = input('Would you like the results to be shown in kN or lbf? ').lower()
print('')
while True:
	if units == 'kn':
		testData = convertToMetric(testData)
		print('Converted Data (kN): ')
		for i in range(len(testData)):
			print(testData[i]['peak load'])
		print('')
		break
	elif units != 'kn' and units != 'lbf':
		print('Invalid unit of measure. Please type kN or lbf.')
		print('')
		units = input('Would you like the results to be shown in kN or lbf? ').lower()
		print('')
	else:
		print('The data has NOT been converted to kN.')
		break

#print units to console
print('')
print('Units: ' + units)

#get average and print to the console
average = getAverage(testData)
print('Average Peak Load: %.2f' % average)  

#Get standard deviation and print stdDev, threeSigma, and confidenceInterval to console
stdDev = getStdDev(testData)
if type(stdDev) == float:
	print('Standard Deviation: %.2f' % stdDev )
	#Get there sigma values and print to console  
	threeSigma = threeSigma(testData)
	print('Lower Three Sigma: %.2f' % threeSigma )
	#Get standard deviations and print to console
	confidenceInterval = confidenceInterval(testData)
	confidenceIntervalValue = getAverage(testData) - (confidenceInterval * getStdDev(testData))
	print('Lower Confidence Interval: %.2f' % confidenceIntervalValue)
	print('')
else:
	stdDev = 'n/a'
	print('Standard Deviation: ' + stdDev)
	threeSigma = 'n/a'
	print('Lower Three Sigma: ' + threeSigma)
	confidenceInterval = 'n/a'
	print('Lower Confidence Interval: ' + confidenceIntervalValue)
	print('')

#Write excel file and save
writeTestDoc()


""" TO DO LIST """

# Display the dataset after each filter is added ??

# overlay it with a histogram????

# Can you add borders for formating the tables? Do we need it? Maybe just fix the cover page......

# Need to fix test document format on record page... it has to do with the way python/openpyxl works with excel

#can just use numpy.average , numpy.stdev, and remove the functions defined above to simplify ???

#use numpy and scipy for now to do all the plotting but look into pandas for other stats plots... lots of options in that module!!!